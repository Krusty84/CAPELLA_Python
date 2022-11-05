# name                 : Report Logical Functions in Excel
# script-type          : Python
# description          : Report Logical Functions in Excel
# Author               : Alexey Sedoykin
# popup                : enableFor(org.polarsys.capella.core.data.la.LogicalFunction)
#
# This script loads the Capella model passed as first argument and list its root LogicalFunction.
# To run it:
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - you can run this script by launching the contextual menu "Run As / EASE Script..." 
#    on this script. 
#    - By default, the model selected is IFE sample (aird path of the model written below)
#  - you can also run this script according to a configuration (script selected, arguments) 
#    and modify the configuration by launching the contextual menu "Run As / Run configurations..." 
#    on this script. 
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "list_logical_functions_in_console.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/List_logical_functions_in_console.py"
#    - define the path to the aird file as first argument in "Script arguments" area: "/In-Flight Entertainment System/In-Flight Entertainment System.aird" (for instance)
# include needed for the Capella modeller API
# for ref https://github.com/eclipse/capella/blob/master/core/plugins/org.polarsys.capella.core.data.helpers/src/org/polarsys/capella/core/data/helpers/cs/delegates/SystemComponentHelper.java
#
#
from openpyxl.styles.builtins import title
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, numbers
from openpyxl.chart import BarChart, Reference

from copy import copy
import os
import re


# Retrieve the Element from the current selection and its aird model path
selected_elem = CapellaElement(CapellaPlatform.getFirstSelectedElement())
aird_path = '/'+ CapellaPlatform.getModelPath(selected_elem)

project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'reports')
xlsx_report = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Logical_functions_report.xlsx'

book = Workbook()

'''
# change this path to execute the script on your model (here is the IFE sample). 
# Uncomment it if you want to use the "Run configuration" instead
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'
'''
'''
#Here is the "Run Configuration" part to uncomment if you want to use this functionality :

#check parameter numbers
if len(argv) != 1:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
'''

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
print("The model name is " + se.get_name())
#
i=0
iDraft=0;
iTBR=0;
iReworkNeed=0;
iTBD=0;
iReviewedOK=0;
iUnderRework=0;
iNonStatus=0;
#
dicDraftFunc={0:{'id':'','name':'','sum':'','outFuncExcNum':'','inpFuncExcNum':'','outFuncExcName':'','inpFuncExcName':''}}
dicTBRFunc={0:{'id':'','name':'','sum':'','outFuncExcNum':'','inpFuncExcNum':'','outFuncExcName':'','inpFuncExcName':''}}
dicReworkNeedFunc={0:{'id':'','name':'','sum':'','outFuncExcNum':'','inpFuncExcNum':'','outFuncExcName':'','inpFuncExcName':''}}
dicTBDFunc={0:{'id':'','name':'','sum':'','outFuncExcNum':'','inpFuncExcNum':'','outFuncExcName':'','inpFuncExcName':''}}
dicReviewedOKFunc={0:{'id':'','name':'','sum':'','outFuncExcNum':'','inpFuncExcNum':'','outFuncExcName':'','inpFuncExcName':''}}
dicUnderReworkFunc={0:{'id':'','name':'','sum':'','outFuncExcNum':'','inpFuncExcNum':'','outFuncExcName':'','inpFuncExcName':''}}
dicNonStatusFunc={0:{'id':'','name':'','sum':'','outFuncExcNum':'','inpFuncExcNum':'','outFuncExcName':'','inpFuncExcName':''}}
#

def funcFilledRawData(rawStrData, iCount):
        rawStrData[iCount] = {}
        rawStrData[iCount]['id'] = lf.get_id()
        rawStrData[iCount]['name'] = lf.get_name()
        rawStrData[iCount]['sum'] = lf.get_summary()
        if(len(lf.get_outgoing())>0):
            rawStrData[iCount]['outFuncExcNum'] = len(lf.get_outgoing())
            print("Out Function Number",len(lf.get_outgoing()))
            for outgoing in lf.get_outgoing():
                print("Out Function Exchange Name",  outgoing.get_name())
                rawStrData[iCount]['outFuncExcName'] = outgoing.get_name()
        else:
            rawStrData[iCount]['outFuncExcNum']='None'
            rawStrData[iCount]['outFuncExcName'] = ''
            
        if(len(lf.get_incoming())>0):
            rawStrData[iCount]['inpFuncExcNum'] = len(lf.get_incoming())
            print("Inp Function Number",len(lf.get_incoming()))
            for incoming in lf.get_incoming():
                print("Inp Function Exchange Name",  incoming.get_name())
                rawStrData[iCount]['inpFuncExcName'] = incoming.get_name()
        else:
            rawStrData[iCount]['inpFuncExcNum']='None'
            rawStrData[iCount]['inpFuncExcName'] = ''
        iCount+=1;
        return iCount
#

def funcCheckFunctionName(functionName):
    if(re.search(r'LogicalFunction \d+$', functionName)is not None):
        return True
    else:
        return False
#
attentionFillYellow=PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')

def funcFilledExcelTable(activeExcelBook, sheetName, sheetSectionName, sheetPosition, columnWidth, rawStrData):
    if(sheetPosition!=0):  
        sheet=activeExcelBook.create_sheet(index=sheetPosition,title=sheetName)
        sheet['A1']="Function ID"
        sheet['B1']="Function Name"
        sheet['C1']="Function Desc"
        sheet['D1']="Outcoming ¹:"
        sheet['E1']="Outcoming Name"
        sheet['F1']="Incoming ¹:"
        sheet['G1']="Incoming Name"
    else:
        sheet=activeExcelBook.create_sheet(title=sheetName)
        sheet['A1']="Function ID"
        sheet['B1']="Function Name"
        sheet['C1']="Function Desc"
        sheet['D1']="Outcoming ¹:"
        sheet['E1']="Outcoming Name"
        sheet['F1']="Incoming ¹:"
        sheet['G1']="Incoming Name"
    
    cellA1=sheet['A1']
    cellB1=sheet['B1']
    cellC1=sheet['C1']
    cellD1=sheet['D1']
    cellE1=sheet['E1']
    cellF1=sheet['F1']
    cellG1=sheet['G1']
        
    sheet.cell(row=1, column=8).value='=HYPERLINK("#Main!A1","GO TO MAIN PAGE")'
    cellLinkToBack=sheet.cell(row=1, column=8)
    cellLinkToBack.font=cellLinkToBack.font.copy(bold=True, italic=False, color='000000')
    cellLinkToBack.alignment = Alignment(horizontal='center')
    cellLinkToBack.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type = "solid")
    sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
            
    cellA1.font=cellA1.font.copy(bold=True, italic=False, color='00b04f')
    cellB1.font=cellB1.font.copy(bold=True, italic=False, color='00b04f') 
    cellC1.font=cellC1.font.copy(bold=True, italic=False, color='00b04f') 
    cellD1.font=cellD1.font.copy(bold=True, italic=False, color='00b04f') 
    cellE1.font=cellE1.font.copy(bold=True, italic=False, color='00b04f') 
    cellF1.font=cellF1.font.copy(bold=True, italic=False, color='00b04f') 
    cellG1.font=cellG1.font.copy(bold=True, italic=False, color='00b04f')
    
    sheet.column_dimensions['A'].width=columnWidth
    sheet.column_dimensions['B'].width=columnWidth
    sheet.column_dimensions['C'].width=columnWidth
    sheet.column_dimensions['D'].width=columnWidth
    sheet.column_dimensions['E'].width=columnWidth
    sheet.column_dimensions['F'].width=columnWidth
    sheet.column_dimensions['G'].width=columnWidth
    sheet.freeze_panes='H2'
    
    if(rawStrData[0]['id']!=""):
                iCountRawStrData=2
                for currentFunc in rawStrData:
                    print(currentFunc, '->', rawStrData[currentFunc])
                    sheet['A'+str(iCountRawStrData)]=rawStrData[currentFunc]['id']
                    if(funcCheckFunctionName(rawStrData[currentFunc]['name'])==True):
                        sheet['B'+str(iCountRawStrData)]=rawStrData[currentFunc]['name']
                        cellB_FuncName=sheet['B'+str(iCountRawStrData)]
                        cellB_FuncName.fill=attentionFillYellow
                    else:
                        sheet['B'+str(iCountRawStrData)]=rawStrData[currentFunc]['name']   
                        sheet['C'+str(iCountRawStrData)]=rawStrData[currentFunc]['sum']
                        sheet['D'+str(iCountRawStrData)]=rawStrData[currentFunc]['outFuncExcNum']
                        sheet['D'+str(iCountRawStrData)].number_format = numbers.FORMAT_NUMBER
                        sheet['E'+str(iCountRawStrData)]=rawStrData[currentFunc]['outFuncExcName']
                        sheet['F'+str(iCountRawStrData)]=rawStrData[currentFunc]['inpFuncExcNum']
                        sheet['F'+str(iCountRawStrData)].number_format = numbers.FORMAT_NUMBER
                        sheet['G'+str(iCountRawStrData)]=rawStrData[currentFunc]['inpFuncExcName']
                    iCountRawStrData+=1
    else:
            sheet.cell(row=2, column=1).value="---"+sheetSectionName+" Functions Not Yet---"
            cellWarning=sheet.cell(row=2, column=1)
            cellWarning.font=cellWarning.font.copy(bold=True, italic=False, color='ff0000')
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        
    for row in sheet.iter_rows():
            for cell in row:
                cell.alignment=cell.alignment.copy(wrapText=True)
#
# print the name of each LogicalFunction
for lf in se.get_all_contents_by_type(LogicalFunction):
    #: :type lf: LogicalFunction
    #print(" - " + lf.get_name())
    if(lf.get_status()=="DRAFT"):
        print("DRAFT - " + lf.get_name())
        iDraft=funcFilledRawData(dicDraftFunc,iDraft)
    elif(lf.get_status()=="TO_BE_REVIEWED"):
        print("TBR - " + lf.get_name())
        iTBR=funcFilledRawData(dicTBRFunc,iTBR)
    elif(lf.get_status()=="REWORK_NECESSARY"):
        print("REWORK_NECESSARY - " + lf.get_name())
        iReworkNeed=funcFilledRawData(dicReworkNeedFunc,iReworkNeed)
    elif(lf.get_status()=="TO_BE_DISCUSSED"):
        print("TBD - " + lf.get_name())
        iTBD=funcFilledRawData(dicTBDFunc,iTBD)
    elif(lf.get_status()=="REVIEWED_OK"):
        print("Reviewed OK - " + lf.get_name())
        iReviewedOK=funcFilledRawData(dicReviewedOKFunc,iReviewedOK)
    elif(lf.get_status()=="UNDER_REWORK"):
        print("Under Rework - " + lf.get_name())
        iUnderRework=funcFilledRawData(dicUnderReworkFunc,iUnderRework)
    else:
        print("Non Status - " + lf.get_name())
        iNonStatus=funcFilledRawData(dicNonStatusFunc,iNonStatus)  
    #sheet['A'+str(i)] = lf.get_name()
    #sheet.cell(row=i, column=1, value=lf.get_name())
    i+=1;

#
print("Non Status Functions #:", iNonStatus);     
print("DRAFT Functions #:", iDraft);    
print("TBR Functions #:", iTBR);
print("REWORK_NECESSARY Functions #:", iReworkNeed);
print("TBD Functions #:", iTBD);
print("Reviewed OK Functions #:", iReviewedOK);
print("Under Rework Functions #:", iUnderRework); 
#

funcFilledExcelTable(book,"Draft_Functions","Draft",0,20,dicDraftFunc)
funcFilledExcelTable(book,"To_Be_Reviewed_Functions","To Be Reviewed",0,20,dicTBRFunc)
funcFilledExcelTable(book,"Rework_Necessary_Functions","Rework Necessary",0,20,dicReworkNeedFunc)
funcFilledExcelTable(book,"To_Be_Discussed_Functions","To Be Discussed",0,20,dicTBDFunc)
funcFilledExcelTable(book,"Reviewed_OK_Functions","Reviewed OK",0,20,dicReviewedOKFunc)
funcFilledExcelTable(book,"Under_Rework_Functions","Under Rework",0,20,dicUnderReworkFunc)
funcFilledExcelTable(book,"Non_Status_Functions","Non Status Rework",1,20,dicNonStatusFunc)
#
firstDummySheet = book['Sheet']
firstDummySheet.title = 'Main'
#
data = [
    ['',''],
    ['Non Status Functions', iNonStatus],
    ['DRAFT Functions', iDraft],
    ['TBR Functions', iTBR],
    ['REWORK_NECESSARY Functions', iReworkNeed],
    ['TBD Functions', iTBD],
    ['Reviewed OK Functions',iReviewedOK],
    ['Under Rework Functions',iUnderRework]
]
#
for row in data:
    firstDummySheet.append(row)
#
label = Reference(firstDummySheet, min_col = 1, min_row = 2, max_row = 8)
data = Reference(firstDummySheet, min_col = 2, min_row = 1, max_row = 8)
#
chart = BarChart()
chart.type = 'bar'
chart.add_data(data, titles_from_data=True)
chart.set_categories(label)
# set the title of the chart
chart.title = "Function statuses"
# set the title of the x-axis
#chart.x_axis.title = "Kind of functions"
# set the title of the y-axis
chart.y_axis.title = "Number of"
# the top-left corner of the chart
firstDummySheet.add_chart(chart,"A1")
#
firstDummySheet.cell(row=1, column=11).value="GO TO..."
cellMenuTitle=firstDummySheet.cell(row=1, column=11)
cellMenuTitle.font=cellMenuTitle.font.copy(bold=True, italic=False, color='000000')
cellMenuTitle.alignment = Alignment(horizontal='center')
cellMenuTitle.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type = "solid")
firstDummySheet.merge_cells(start_row=1, start_column=11, end_row=1, end_column=16)
#
firstDummySheet.cell(row=2, column=11).value = '=HYPERLINK("#Non_Status_Functions!A1","Non Status Functions")'
firstDummySheet.merge_cells(start_row=2, start_column=11, end_row=2, end_column=16)
firstDummySheet.cell(row=3, column=11).value = '=HYPERLINK("#Draft_Functions!A1","Draft Functions")'
firstDummySheet.merge_cells(start_row=3, start_column=11, end_row=3, end_column=16)
firstDummySheet.cell(row=4, column=11).value = '=HYPERLINK("#To_Be_Reviewed_Functions!A1","To Be Reviewed Functions")'
firstDummySheet.merge_cells(start_row=4, start_column=11, end_row=4, end_column=16)
firstDummySheet.cell(row=5, column=11).value = '=HYPERLINK("#Rework_Necessary_Functions!A1","Rework Necessary Functions")'
firstDummySheet.merge_cells(start_row=5, start_column=11, end_row=5, end_column=16)
firstDummySheet.cell(row=6, column=11).value = '=HYPERLINK("#To_Be_Discussed_Functions!A1","To Be Discussed Functions")'
firstDummySheet.merge_cells(start_row=6, start_column=11, end_row=6, end_column=16)
firstDummySheet.cell(row=7, column=11).value = '=HYPERLINK("#Reviewed_OK_Functions!A1","Reviewed OK Functions")'
firstDummySheet.merge_cells(start_row=7, start_column=11, end_row=7, end_column=16)
firstDummySheet.cell(row=8, column=11).value = '=HYPERLINK("#Under_Rework_Functions!A1","Under Rework Functions")'
firstDummySheet.merge_cells(start_row=8, start_column=11, end_row=8, end_column=16)
#
book.save(xlsx_report)
CapellaPlatform.refresh(folder)
os.system("start EXCEL.exe "+xlsx_report)